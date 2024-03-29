from PyQt6 import QtWidgets, QtCore, uic, QtGui
from PyQt6.QtCore import Qt, QAbstractTableModel, pyqtSignal
from PyQt6.QtWidgets import QMessageBox, QListWidgetItem, QTableView, QApplication, QFileDialog
from PyQt6.QtSql import QSqlDatabase, QSqlQuery, QSqlTableModel
from PyQt6.QtGui import QColor, QStandardItemModel
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from decimal import Decimal
import sys, csv, math, webbrowser, os, traceback
'''
To-do
Talk with the worst and see if I can just directly hook into savy API for frequent updates somehow on players?
ALL - Tooltips?
MAIN - build a full readme to walkthrough how to operate all windows
TOURNY - Division A flag to assume 4 star battles for 2 fights every day for est calculator
       - add a way to check most recent tournament targets and what day they were done on
IMPORT - 
FIGHTS - 
CTC - crew training calculator: error checking and prevention of invalid values
'''
def create_connection():
      databases = {
            "targetdb": "targets.db",
            "tournydb": "tournyfights.db",
            "legendsdb": "legendfights.db",
            "pvpdb": "pvpfights.db"
      }
      for name, filename in databases.items():
            db = QSqlDatabase.addDatabase('QSQLITE', name)
            db.setDatabaseName(filename)
            if not db.open():
                  throwErrorMessage("Fatal Error", f"{name.capitalize()} DB did not open properly")
                  return False
      return True
def create_table():
      targetsQuery = QSqlQuery(QSqlDatabase.database("targetdb"))
      tournyQuery = QSqlQuery(QSqlDatabase.database("tournydb"))
      legendQuery = QSqlQuery(QSqlDatabase.database("legendsdb"))
      pvpQuery = QSqlQuery(QSqlDatabase.database("pvpdb"))
      targetsQuery.exec("CREATE TABLE IF NOT EXISTS players (playername TEXT PRIMARY KEY, fleetname TEXT NOT NULL, laststars TEXT NOT NULL, beststars TEXT NOT NULL, trophies TEXT NOT NULL, maxtrophies TEXT NOT NULL, notes TEXT NOT NULL)")
      tournyQuery.exec("CREATE TABLE IF NOT EXISTS fights (name TEXT NOT NULL, rewards TEXT NOT NULL, datetag TEXT NOT NULL, hpremain INTEGER NOT NULL, winloss TEXT NOT NULL, UNIQUE(name, rewards, datetag, hpremain, winloss))")
      legendQuery.exec("CREATE TABLE IF NOT EXISTS fights (name TEXT NOT NULL, rewards TEXT NOT NULL, datetag TEXT NOT NULL, hpremain INTEGER NOT NULL, winloss TEXT NOT NULL, UNIQUE(name, rewards, datetag, hpremain, winloss))")
      pvpQuery.exec("CREATE TABLE IF NOT EXISTS fights (name TEXT NOT NULL, rewards TEXT NOT NULL, datetag TEXT NOT NULL, hpremain INTEGER NOT NULL, winloss TEXT NOT NULL, UNIQUE(name, rewards, datetag, hpremain, winloss))")
def write_to_fights_database(data, fights):
      if fights == "tourny":
            query = QSqlQuery(QSqlDatabase.database("tournydb"))
      if fights == "legends":
            query = QSqlQuery(QSqlDatabase.database("legendsdb"))
      if fights == "pvp":
            query = QSqlQuery(QSqlDatabase.database("pvpdb"))
      query.prepare("INSERT OR REPLACE INTO fights(name, rewards, datetag, hpremain, winloss) VALUES(?, ?, ?, ?, ?)")
      for i in range(5):
            query.bindValue(i, data[i].strip())
      if not query.exec():
            throwErrorMessage("FightsDB [write_to_fights_database]:", query.lastError().text())
            return False
      return True
def write_to_targets_database(data):
      query = QSqlQuery(QSqlDatabase.database("targetdb"))
      query.prepare("INSERT OR REPLACE INTO players(playername, fleetname, laststars, beststars, trophies, maxtrophies, notes) VALUES(?, ?, ?, ?, ?, ?, ?)")
      for i in range(7):
            query.bindValue(i, data[i])
      if not query.exec():
            throwErrorMessage("targetdb [write_to_targets_database]:", query.lastError().text())
            return False
      return True
def throwErrorMessage(text, dump):
      msg = QMessageBox()
      msg.setIcon(QMessageBox.Icon.Critical)
      errortext = "Error: "+text
      msg.setText(errortext)
      msg.setInformativeText(dump)
      msg.setWindowTitle("Error")
      msg.exec()
class MainWindow(QtWidgets.QMainWindow):
      def __init__(self):
            super(MainWindow, self).__init__() 
            uic.loadUi(os.path.join('_internal','pss-ttt.ui'), self)
            self.show()
            
            self.lockUnlockButton.clicked.connect(self.changeButtonText)
            self.searchButton.clicked.connect(self.searchPlayer)
            self.saveNewData.clicked.connect(self.submitNewPlayerData)
            self.resetButton.clicked.connect(self.resetDataFields)
            self.pixyshipLayoutButton.clicked.connect(self.pixyshipURL)
            self.submitNewDataButton.clicked.connect(self.open_fightDialog)
            self.importDialogButton.clicked.connect(self.open_importDialog)
            self.playerBrowserSearchButton.clicked.connect(self.open_playerBrowser)
            self.fleetSearchButton.clicked.connect(self.open_fleetBrowser)
            self.fleetBrowserSearchButton.clicked.connect(self.open_fleetNameBrowser)
            self.delLegendsButton.clicked.connect(lambda: self.deleteSelectedLine("legends", self.legendsTable))
            self.delTournyButton.clicked.connect(lambda: self.deleteSelectedLine("tourny", self.tournyTable))
            self.delPVPButton.clicked.connect(lambda: self.deleteSelectedLine("pvp", self.pvpTable))
            self.tournyStarsWindow.clicked.connect(self.open_tournamentStarsCalc)
            self.crewTrainerButton.clicked.connect(self.open_crewTrainer)
            self.starTargetTrackButton.clicked.connect(self.open_stt)
            self.exportFightsButton.clicked.connect(self.exportFightsToCSV)
            
            self.fightDialog = FightDataConfirmation(parent=self)
            self.fightDialog.fightDataSaved.connect(self.receiveFightData)

            '''self.fleetBrowser = FleetDialogBox()
            self.fleetBrowser.copyFleetSearchClicked.connect(self.handleCopyFleetSearchClicked)
            
            self.fleetNameBrowser = FleetNameDialogBox()
            self.fleetNameBrowser.copyFleetNameSearchClicked.connect(self.handleCopyFleetNameSearchClicked)

            self.playerBrowser = PlayerDialogBox()
            self.playerBrowser.copyPlayerSearchClicked.connect(self.handleCopyPlayerSearchClicked)'''

            self.player_dialog = FilteredListDialog("Player Dialog", "Player Name", "Search Player")
            self.player_dialog.copyItemSearchClicked.connect(self.handleCopyPlayerSearchClicked)
            self.fleet_dialog = FilteredListDialog("Fleet Dialog", "Fleet Name", "Search Fleet")
            self.fleet_dialog.copyItemSearchClicked.connect(self.handleCopyFleetSearchClicked)
            self.fleet_name_dialog = FilteredListDialog("Fleet Name Dialog", "Fleet Name", "Search Fleet Name")
            self.fleet_name_dialog.copyItemSearchClicked.connect(self.handleCopyFleetNameSearchClicked)

            self.starCalculator = TournamentDialogBox()

            self.importDialog = ImportDialogBox()

            self.trainingDialog = CrewTrainerDialogBox()

            self.starTargetTrack = StarTargetTrackDialogBox(parent=self)
            
            tournyQuery = QSqlQuery(QSqlDatabase.database("tournydb"))
            legendQuery = QSqlQuery(QSqlDatabase.database("legendsdb"))
            pvpQuery = QSqlQuery(QSqlDatabase.database("pvpdb"))
            self.update_SQL(tournyQuery, "Tourny")
            self.update_SQL(legendQuery, "Legend")
            self.update_SQL(pvpQuery, "PVP")
      def update_SQL(self, query, db_name):
            alter_queries = [
                  "ALTER TABLE fights ADD COLUMN hpremain INTEGER DEFAULT 0",
                  "ALTER TABLE fights ADD COLUMN winloss TEXT DEFAULT 'Draw'"
                  ]
            # Define UPDATE query to set default values
            update_query = (
                  "UPDATE fights "
                  "SET hpremain = 0, winloss = 'Draw' "
                  "WHERE hpremain IS NULL OR winloss IS NULL"
                  )
            # Execute ALTER TABLE queries
            for alter_query in alter_queries:
                  if not query.exec(alter_query):
                        if not query.lastError().text().startswith("duplicate"):
                              print(f"Error: {db_name} Alter")
                              print("Error:", query.lastError().text())
            # Execute UPDATE query
            if not query.exec(update_query):
                  if not query.lastError().text().startswith("duplicate"):
                        print(f"Error: {db_name} Update")
                        print("Error:", query.lastError().text())
      def open_fightDialog(self):
            self.fightDialog.exec()
      def open_playerBrowser(self):
            self.player_dialog.populateList("SELECT playername FROM players", None)
            self.player_dialog.exec()
      def open_fleetBrowser(self):
            self.fleet_dialog.populateList("SELECT playername FROM players WHERE fleetname = :fleet_name", self.fleetName.toPlainText())
            self.fleet_dialog.exec()
      def open_fleetNameBrowser(self):
            self.fleet_name_dialog.populateList("SELECT fleetname FROM players", None)
            self.fleet_name_dialog.exec()
      def open_importDialog(self):
            self.importDialog.importDialogLabel.setText("")
            self.importDialog.importFilenameBox.setPlainText("")
            self.importDialog.exec()
      def open_tournamentStarsCalc(self):
            self.starCalculator.exec()
      def open_crewTrainer(self):
            self.trainingDialog.exec()
      def open_stt(self):
            self.starTargetTrack.populateSTT(self.playerNameSearchBox.toPlainText())
            self.starTargetTrack.exec()
      def exportFightsToCSV(self):
            filepath = os.path.join('_internal', 'exportedfights.csv')
            fights_data = []
            for db_name in ("tournydb", "legendsdb", "pvpdb"):
                  db = QSqlDatabase.database(db_name)
                  query = QSqlQuery(db)
                  query.prepare("SELECT name, rewards, datetag, hpremain, winloss FROM fights ORDER BY name")
                  if not query.exec():
                        throwErrorMessage("Query execution failed [exportFightsToCSV]: ", query.lastError().text())
                        return
                  while query.next():
                        name = query.value(0)
                        rewards = query.value(1)
                        datetag = query.value(2)
                        hpremain = query.value(3)
                        winloss = query.value(4)
                        fights_data.append((db_name, name, rewards, datetag, hpremain, winloss))

            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                  writer = csv.writer(csvfile)
                  writer.writerow(["Database", "Name", "Rewards", "Datetag", "HP Remaining", "Win/Loss"])
                  for row in fights_data:
                        writer.writerow(row)
            self.exportPlayerDataToCSV()
      def exportPlayerDataToCSV(self):
            filepath = os.path.join('_internal', 'exportedplayers.csv')
            player_data = []
            db = QSqlDatabase.database("targetdb")
            query = QSqlQuery(db)
            query.prepare("SELECT playername, fleetname, laststars, beststars, trophies, maxtrophies, notes FROM players ORDER BY playername")
            if not query.exec():
                  throwErrorMessage("Query Execution failed [exportPlayerDataToCSV]: ", query.lastError().text())
                  return
            while query.next():
                  playername = query.value(0)
                  fleetname = query.value(1)
                  laststars = query.value(2)
                  beststars = query.value(3)
                  trophies = query.value(4)
                  maxtrophies = query.value(5)
                  notes = query.value(6)
                  player_data.append((playername, fleetname, laststars, beststars, trophies, maxtrophies, notes))
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                  writer = csv.writer(csvfile)
                  writer.writerow(["Player Name", "Fleet Name", "Last Stars", "Best Stars", "Trophies", "Max Trophies", "Notes"])
                  for row in player_data:
                        writer.writerow(row)
      def handleCopyFleetSearchClicked(self, selected_text, close_dialog):
            self.resetDataFields()
            self.playerNameSearchBox.setPlainText(selected_text)
            self.searchPlayer()
            if close_dialog:
                  self.fleet_dialog.accept()
      def handleCopyFleetNameSearchClicked(self, selected_text, close_dialog):
            self.resetDataFields()
            self.fleetName.setPlainText(selected_text)
            if close_dialog:
                  self.fleet_name_dialog.accept()
      def handleCopyPlayerSearchClicked(self, selected_text, close_dialog):
            self.resetDataFields()
            self.playerNameSearchBox.setPlainText(selected_text)
            self.searchPlayer()
            if close_dialog:
                  self.player_dialog.accept()
      def pixyshipURL(self):
            searchName = self.playerNameSearchBox.toPlainText()
            if searchName == "":
                  return
            else:
                  webpage = "https://pixyship.com/players?player="+searchName
                  return webbrowser.open(webpage)         
      def resetDataFields(self):
            self.playerNameSearchBox.clear()
            self.fleetName.clear()
            self.lastStars.clear()
            self.bestStars.clear()
            self.currentTrophies.clear()
            self.maxTrophies.clear()
            self.playerNotes.clear()
            model = QSqlTableModel()
            model.clear()
            self.tournyTable.setModel(model)
            self.tournyTable.show()
            self.legendsTable.setModel(model)
            self.legendsTable.show()
            self.pvpTable.setModel(model)
            self.pvpTable.show()
      def searchPlayer(self):
            text = self.playerNameSearchBox.toPlainText().strip()
            if not text:
                  return
            self.resetDataFields()
            self.playerNameSearchBox.setPlainText(text)
            query = QSqlQuery("SELECT * FROM players", QSqlDatabase.database("targetdb"))
            query.prepare("SELECT * FROM players WHERE playername = :text")
            query.bindValue(":text", text)
            if query.exec():
                  if query.next():
                        self.fleetName.setPlainText(query.value(1))
                        self.lastStars.setPlainText(query.value(2))
                        self.bestStars.setPlainText(query.value(3))
                        self.currentTrophies.setPlainText(query.value(4))
                        self.maxTrophies.setPlainText(query.value(5))
                        self.playerNotes.setPlainText(query.value(6))
                        self.updateFightTables("tourny")
                        self.updateFightTables("legends")
                        self.updateFightTables("pvp")
                  else:
                        return
            else:
                  throwErrorMessage("Query execution failed [searchPlayer]: ", query.lastError().text())
                  return                 
      def changeButtonText(self):
            is_locked = self.lockUnlockButton.text() == "Locked"
            self.lockUnlockButton.setText("Unlocked" if is_locked else "Locked")
            self.playerDataChange(not is_locked)
      def playerDataChange(self, boolean_value):
            self.lastStars.setReadOnly(boolean_value)
            self.bestStars.setReadOnly(boolean_value)
            self.currentTrophies.setReadOnly(boolean_value)
            self.maxTrophies.setReadOnly(boolean_value)
            self.playerNotes.setReadOnly(boolean_value)     
      def submitNewPlayerData(self):
            try:
                  last_stars = int(self.lastStars.toPlainText())
            except ValueError:
                  last_stars = 0
            try:
                  best_stars = int(self.bestStars.toPlainText())
            except ValueError:
                  best_stars = 0
            try:
                  current_trophies = int(self.currentTrophies.toPlainText())
            except ValueError:
                  current_trophies = 0
            try:
                  max_trophies = int(self.maxTrophies.toPlainText())
            except ValueError:
                  max_trophies = 0
            if last_stars > best_stars:
                  best_stars = last_stars
            if current_trophies > max_trophies:
                  max_trophies = current_trophies
            player_data = [self.playerNameSearchBox.toPlainText(), self.fleetName.toPlainText(), last_stars,
                        best_stars, current_trophies, max_trophies, self.playerNotes.toPlainText()]
            if write_to_targets_database(player_data):
                  player_data.clear()
            else:
                  throwErrorMessage("targetdb: Error writing data to the database - Dumping data [submitNewPlayerData]", player_data)
      def updateFightTables(self, fights):
            name = self.playerNameSearchBox.toPlainText()
            query_error_message = f"{fights.capitalize()}DB Update: "

            if fights in ("tourny", "legends", "pvp"):
                  db_name = f"{fights}db"
                  table_widget = getattr(self, f"{fights}Table")
                  query = QSqlQuery(QSqlDatabase.database(db_name))
                  query.prepare("SELECT rewards, datetag, hpremain, winloss FROM fights WHERE name LIKE ?")
                  if query.lastError().isValid():
                        throwErrorMessage(query_error_message&" [updateFightTables]", query.lastError().text())
                        sys.exit(-1)
                  query.addBindValue(f'%{name}%')
                  query.exec()

                  model = self.CustomSqlTableModel()
                  model.setQuery(query)
                  table_widget.setModel(model)
                  table_widget.setColumnWidth(0,40)
                  table_widget.setColumnWidth(1,70)
                  table_widget.setColumnWidth(2,25)
                  table_widget.setColumnWidth(3,40)
                  table_widget.show()
            else:
                  throwErrorMessage("Fights Table Error [updateFightTables]", "Invalid fights type")
      def receiveFightData(self, rewards, remainhp, result, fight):
            self.submitFightData(rewards, remainhp, result, fight)
      def submitFightData(self, rewards, remainhp, result, fight):
            todaysdate = str(date.today())
            fight_data = [self.playerNameSearchBox.toPlainText(), str(rewards), todaysdate, str(remainhp), result]
            if not write_to_fights_database(fight_data, fight):
                  print(fight_data, " | ", fight)
                  throwErrorMessage("FightsDB: Error writing data to the database - Dumping data [submitFightData]", fight_data)
            getattr(self, f"{fight}Table").clearSpans()
            self.updateFightTables(fight) 
            fight_data.clear()
      def deleteSelectedLine(self, table_name, tableView):
            selected_indexes = tableView.selectionModel().selectedRows()
            db = QSqlDatabase.database(f"{table_name}db")
            if not selected_indexes:
                  return

            name_data = []
            rewards_data = []
            datetag_data = []
            hpremain_data = []
            winloss_data = []

            for index in selected_indexes:
                  rewards = tableView.model().index(index.row(), 0).data()
                  datetag = tableView.model().index(index.row(), 1).data()
                  hpremain = tableView.model().index(index.row(), 2).data()
                  winloss = tableView.model().index(index.row(), 3).data()
                  rewards_data.append(rewards)
                  datetag_data.append(datetag)
                  hpremain_data.append(hpremain)
                  winloss_data.append(winloss)

            name_data.append(self.playerNameSearchBox.toPlainText())

            query = QSqlQuery(db)
            sql_query = "DELETE FROM fights WHERE name IN ({}) AND rewards IN ({}) AND datetag IN ({}) AND hpremain IN ({}) AND winloss IN ({})".format(
                  ", ".join(["'{}'".format(str(name)) for name in name_data]),  
                  ", ".join(["'{}'".format(str(data)) for data in rewards_data]),
                  ", ".join(["'{}'".format(str(data)) for data in datetag_data]),
                  ", ".join(["'{}'".format(float(data)) for data in hpremain_data]),
                  ", ".join(["'{}'".format(str(data)) for data in winloss_data]))

            if not query.exec(sql_query):
                  throwErrorMessage("Record Deletion Error : [deleteSelectedLine]["+table_name+"]", query.lastError().text())
                  return

            tableView.model().clear()
            self.updateFightTables(table_name)
      class CustomSqlTableModel(QSqlTableModel):
            def __init__(self, parent=None):
                  super().__init__(parent)
            def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
                  if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
                        headers = {
                              0: "★/🏆",
                              1: "Date",
                              2: "HP",
                              3: "W/L/D"
                              }
                        return headers.get(section, super().headerData(section, orientation, role))
                  return super().headerData(section, orientation, role)
            def data(self, index, role):
                  if not index.isValid():
                        return None
                  if role == Qt.ItemDataRole.BackgroundRole:
                        column_name = "datetag"  # Replace this with the actual column name
                        datetag_value = self.record(index.row()).value(column_name)
                        datetag_date = datetime.strptime(datetag_value, "%Y-%m-%d")
                        last_day_of_month = datetime(datetag_date.year, datetag_date.month, 1) + timedelta(days=32)
                        last_day_of_month = last_day_of_month.replace(day=1) - timedelta(days=1)
                        if (last_day_of_month - datetag_date).days <= 7:
                              return QColor(181,214,232)
                  return super().data(index, role)
class FilteredListDialog(QtWidgets.QDialog):
      copyItemSearchClicked = QtCore.pyqtSignal(str, bool)
      def __init__(self, title, filter_label, filter_placeholder, parent=None):
            super().__init__(parent)
            uic.loadUi(os.path.join('_internal', 'pss-ttt-fld.ui'), self)
            self.setWindowTitle(title)

            self.itemSearchClose.clicked.connect(self.accept)
            self.copyItemSearch.clicked.connect(self.onCopyItemSearchClicked)
            self.filterBox.textChanged.connect(self.filterList)
            self.charLimiter.valueChanged.connect(self.filterListLength)
            self.charLimiterCheckBox.stateChanged.connect(self.filterListLength)

            self.filterLabel.setText(filter_label)
            self.filterBox.setPlaceholderText(filter_placeholder)
      def populateList(self, query_string, query_args):
            self.itemList.clear()
            if not QSqlDatabase.database("targetdb").isOpen():
                  if not QSqlDatabase.database("targetdb").open():
                        throwErrorMessage("Database Connection Failure [populateList]", "Targets DB did not open properly")
                        return
            
            query = QSqlQuery(QSqlDatabase.database("targetdb"))
            query.prepare(query_string)
            if query_args:
                  query.bindValue(":fleet_name", query_args)

            if not query.exec():
                  throwErrorMessage("Query Execution Failure [populateList]", query.lastError().text())
                  return
            fleet_names = set()
            while query.next():
                  item_text = query.value(0)
                  if item_text not in fleet_names:
                        fleet_names.add(item_text)
                        item = QListWidgetItem(item_text)
                        self.itemList.addItem(item)
      def onCopyItemSearchClicked(self):
            selected_item = self.itemList.currentItem()
            if selected_item:
                  selected_text = selected_item.text()
                  self.copyItemSearchClicked.emit(selected_text, True)
      def filterList(self):
            filter_text = self.filterBox.toPlainText().lower()
            for i in range(self.itemList.count()):
                  item = self.itemList.item(i)
                  item_text = item.text().lower()
                  if filter_text in item_text:
                        item.setHidden(False)
                  else:
                        item.setHidden(True)
      def filterListLength(self):
            if self.charLimiterCheckBox.isChecked():
                  length_threshold = self.charLimiter.value()
                  for i in range(self.itemList.count()):
                        item = self.itemList.item(i)
                        item_text = item.text()
                        if len(item_text) == length_threshold:
                              item.setHidden(False)
                        else:
                              item.setHidden(True)
class FightDataConfirmation(QtWidgets.QDialog):
      fightDataSaved = pyqtSignal(int, float, str, str)
      def __init__(self, parent=None):
            super().__init__(parent)
            uic.loadUi(os.path.join('_internal','pss-ttt-fdb.ui'), self)

            self.submitFightDataButton.clicked.connect(self.saveFightData)
      def saveFightData(self):
            rewards = int(self.fightRewardBox.toPlainText())
            remainhp = float(self.fightHPBox.toPlainText())
            result = str(self.fightResultBox.currentText())
            fight = str(self.fightTypeBox.currentText())
            self.fightDataSaved.emit(rewards, remainhp, result, fight)
            self.accept()
class ImportDialogBox(QtWidgets.QDialog):
      progress_signal = pyqtSignal(int)
      counter = 0
      max_counter = 0
      def __init__(self):
            super().__init__()
            uic.loadUi(os.path.join('_internal','pss-ttt-importdialog.ui'), self)
            self.importTargetsButton.clicked.connect(self.import_data)
            self.importBrowse.clicked.connect(self.open_fileBrowser)

            self.progress_signal.connect(self.updateProgressBar, QtCore.Qt.ConnectionType.DirectConnection)
      def open_fileBrowser(self):
            options = QFileDialog.Option.DontUseNativeDialog
        
            # Display the file dialog
            selected_files, _ = QFileDialog.getOpenFileNames(self, "Select File", "", "All Files (*)", options=options)
        
            # Copy the selected file paths to the QPlainTextEdit widget
            for file_path in selected_files:
                  self.importFilenameBox.appendPlainText(file_path)
      def import_data(self):
            self.importDialogLabel.setText("Data importing has begun.<br>Please note the application make appear to freeze<br>It is not frozen and just processing in the background<br>This will update once finished")
            QApplication.processEvents()
            file_path = self.importFilenameBox.toPlainText().strip()
            
            if not file_path:
                  throwErrorMessage("Import Error", "No file provided or name is incorrect")
                  return
            if file_path.lower().endswith('.csv'):
                  self.max_counter = total_targets_imported = self.importCSV(file_path)
                  file_type = "CSV"
            elif file_path.lower().endswith(('.xls', '.xlsx')):
                  self.max_counter = total_targets_imported = self.importExcel(file_path)
                  file_type = "Excel"
            else:
                  throwErrorMessage("Unsupported file format", "Only able to accept manicured excel formats or csv from Dolores 2.0 bot")
                  return
            self.importDialogLabel.setText(f"Total Targets Import ({file_type}): {total_targets_imported} targets")
            for i in range(self.max_counter):
                  self.progress_signal.emit(self.counter)
      def updateProgressBar(self, value):
            self.importProgressBar.setValue(int((value / self.max_counter) * 100))
      def importExcel(self, file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            if not QSqlDatabase.database("targetdb").open():
                  throwErrorMessage("Targets Database Error", "Unable to open targets database for import")
                  return self.counter
            query = QSqlQuery(QSqlDatabase.database("targetdb"))
            query.exec(f"CREATE TABLE IF NOT EXISTS players (playername TEXT, fleetname TEXT, laststars INT, beststars INT, trophies INT, maxtrophies INT, notes TEXT)")

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row assuming the first row is header
                  playername, fleetname, laststars, beststars, trophies, maxtrophies, notes = row
                  query.prepare(f"INSERT OR REPLACE INTO players (playername, fleetname, laststars, beststars, trophies, maxtrophies, notes) "
                      "VALUES (?, ?, ?, ?, ?, ?, ?)")
                  query.addBindValue(playername)
                  query.addBindValue(fleetname)
                  query.addBindValue(laststars)
                  query.addBindValue(beststars)
                  query.addBindValue(trophies)
                  query.addBindValue(maxtrophies)
                  query.addBindValue(notes)
                  if query.exec():
                        self.counter += 1
                        self.progress_signal.emit(self.counter)
                  else:
                        throwErrorMessage("Targets Database Insertion Error:", query.lastError().text())
            
            QSqlDatabase.database("targetdb").close()
            return self.counter
      def importCSV(self, file_path):
            data_to_insert = []
            with open(file_path, 'r', encoding='utf-8') as csvfile:
                  reader = csv.reader(csvfile, delimiter=';')
                  next(reader)
                  for row in reader:
                        #specific_data = (row[2], row[1], row[7], row[7], row[5], row[6], " ")
                        #data_to_insert.append(specific_data)
                        playername, fleetname, laststars, beststars, trophies, maxtrophies, notes = row[2], row[1], row[7], row[7], row[5], row[6], ' '
                        laststars_int = int(laststars)
                        beststars_int = int(beststars)
                        notes_str = notes
            
                        db = QSqlDatabase.database("targetdb")
                        query = QSqlQuery(db)
                        query.prepare("SELECT laststars, beststars, notes FROM players WHERE playername = ? AND fleetname = ?")
                        query.addBindValue(playername)
                        query.addBindValue(fleetname)
                        query.exec()
                        if query.next():
                              laststars_db = int(query.value(0))
                              beststars_db = int(query.value(1))
                              notes_db = str(query.value(2))
                              if laststars_int < laststars_db: #overwrites old star data
                                    laststars_int = laststars_db
                              if beststars_int < beststars_db: # overwrites old star data
                                    beststars_int = beststars_db
                              if notes_db != ' ': # prevent overwriting of notes *facepalm*
                                    notes_str = notes_db
                        data_to_insert.append((playername, fleetname, laststars_int, beststars_int, trophies, maxtrophies, notes_str))
    
            db = QSqlDatabase.database("targetdb")
            if not db.open():
                  throwErrorMessage("Targets Database Error", "Unable to open targets database for import")
                  return self.counter
    
            query = QSqlQuery(db)
            query.exec("CREATE TABLE IF NOT EXISTS players (playername TEXT PRIMARY KEY, fleetname TEXT NOT NULL, laststars TEXT NOT NULL, beststars TEXT NOT NULL, trophies TEXT NOT NULL, maxtrophies TEXT NOT NULL, notes TEXT NOT NULL)")
    
            self.max_counter = len(data_to_insert)
            for specific_data in data_to_insert:
                  playername, fleetname, laststars, beststars, trophies, maxtrophies, notes = specific_data
                  count_query = QSqlQuery(db)
                  count_query.prepare("SELECT COUNT(*) FROM players WHERE playername = ? AND fleetname = ?")
                  count_query.addBindValue(playername)
                  count_query.addBindValue(fleetname)
                  count_query.exec()
                  count_query.next()
                  count = count_query.value(0)
                  
                  query.prepare("INSERT OR REPLACE INTO players (playername, fleetname, laststars, beststars, trophies, maxtrophies, notes) VALUES (?, ?, ?, ?, ?, ?, ?)")
                  query.addBindValue(playername)
                  query.addBindValue(fleetname)
                  query.addBindValue(laststars)
                  query.addBindValue(beststars)
                  query.addBindValue(trophies)
                  query.addBindValue(maxtrophies)
                  query.addBindValue(notes)
        
                  if query.exec():
                        self.counter += 1
                        self.progress_signal.emit(self.counter)
                  else:
                        throwErrorMessage("Database Error", query.lastError().text())
            db.commit()
            return self.counter
class TournamentDialogBox(QtWidgets.QDialog):
      def __init__(self):
            super().__init__()
            uic.loadUi(os.path.join('_internal','pss-ttt-tsc.ui'), self)
            
            self.starTableHeaders = ['Star Goal', 'Fight 1', 'Fight 2', 'Fight 3', 'Fight 4', 'Fight 5', 'Fight 6']
            self.starsTable = [
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],]
            self.loadStarsTableFromCSV(os.path.join('_internal','starstable.csv'))
            self.tournamentTable = self.findChild(QTableView, "starsTableView")
            self.model = self.TournamentTableModel(self.starsTable, self)
            self.tournamentTable.setModel(self.model)
            for i in range(7):
                  self.tournamentTable.setColumnWidth(i,50)
            self.calculateStars.clicked.connect(self.calculateStarsGoal)     
            self.resetStarsTableButton.clicked.connect(self.resetStarsTable)
      def updateActualStarsBox(self):
            total_sum = sum(self.model.getCellValue(7, col) for col in range(self.model.columnCount(None)))
            self.actualStarsBox.setPlainText(str(total_sum))
            self.saveStarsTableToCSV(os.path.join('_internal','starstable.csv'))
      def saveStarsTableToCSV(self, filename):
            with open(filename, 'w', newline='') as csvfile:
                  writer = csv.writer(csvfile)
                  writer.writerows(self.starsTable)
      def loadStarsTableFromCSV(self, filename):
            if os.path.exists(filename):
                  with open(filename, newline='') as csvfile:
                        reader = csv.reader(csvfile)
                        self.starsTable = [list(map(int, row)) for row in reader]
            else:
                  self.starsTable = [
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],]
                  self.saveStarsTableToCSV(filename)
      def resetStarsTable(self):
            self.starsTable = [
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],
                  [0,0,0,0,0,0,0],]
            self.saveStarsTableToCSV(os.path.join('_internal','starstable.csv'))
            self.model = self.TournamentTableModel(self.starsTable, self)
            self.tournamentTable.setModel(self.model)
      class starsErrorDialog(QtWidgets.QDialog):
            def __init__(self):
                  super().__init__()
            def checkStarsErrors(self):
                  winsPerDay = self.winsPerDayBox.toPlainText().strip()
                  myTrophies = self.myTrophiesBox.toPlainText().strip()
                  strengthRatio = self.strengthRatioBox.toPlainText().strip()

                  if not winsPerDay or not myTrophies or not strengthRatio:
                        error_message = QMessageBox()
                        error_message.setIcon(QMessageBox.Icon.Warning)
                        error_message.setText("Error: Please fill out all 3 boxes for star calculations.")
                        error_message.setWindowTitle("Missing Information")
                        error_message.exec()
                        return False
                  else:
                        return True
      class TournamentTableModel(QAbstractTableModel):
            def __init__(self, data, parent=None):
                  super().__init__()
                  self._data = data
                  self.parent = parent
                  self.verticalHeaders = ['Star Goal', 'Fight 1', 'Fight 2', 'Fight 3', 'Fight 4', 'Fight 5', 'Fight 6', 'Stars Gained']
                  self.horizontalHeaders = ['Day 1', 'Day 2', 'Day 3', 'Day 4', 'Day 5', 'Day 6', 'Day 7']
            def rowCount(self, index):
                  return len(self._data)
            def columnCount(self, index):
                  return len(self._data[0])
            def data(self, index, role=Qt.ItemDataRole.DisplayRole):
                  if index.isValid() and role == Qt.ItemDataRole.DisplayRole:
                        return str(self._data[index.row()][index.column()])
            def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
                  if role == Qt.ItemDataRole.EditRole and index.isValid():
                        try:
                              int_value = int(value)
                              self._data[index.row()][index.column()] = int_value
                        except ValueError:
                              self._data[index.row()][index.column()] = 0
                        self.dataChanged.emit(index, index)
                        for col in range(self.columnCount(None)):
                              try:
                                    self._data[7][col] = sum(int(self._data[i][col]) for i in range(1,7))
                              except ValueError:
                                    self._data[7][col] = 0
                        self.parent.updateActualStarsBox()
                        return True
                  return False
            def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
                  if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
                        return self.horizontalHeaders[section]
                  if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Vertical:
                        return self.verticalHeaders[section]
            def getCellValue(self, row, column):
                  return self._data[row][column]
            def flags(self, index):
                  if index.row() == 0 or index.row() == 7:
                        return Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
                  else:
                        return Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable
      def calculateStarsGoal(self):
            if not self.starsErrorDialog.checkStarsErrors(self):
                  return
                        
            start_value = Decimal(0)
            winsPerDay = Decimal(self.winsPerDayBox.toPlainText())
            myTrophies = Decimal(self.myTrophiesBox.toPlainText())
            strengthRatio = Decimal(self.strengthRatioBox.toPlainText())
            end_value = Decimal(0)
            todaysStarsTarget = Decimal(0)
            for i in range(7):
                  start_value = end_value
                  trophyStarsTarget = math.floor((myTrophies*strengthRatio)/1000)
                  starsStarsTarget = math.floor(Decimal(start_value*Decimal(".15")*strengthRatio))
                  if trophyStarsTarget > starsStarsTarget:
                        todaysStarsTarget = trophyStarsTarget
                  else:
                        todaysStarsTarget = starsStarsTarget
                  end_value = todaysStarsTarget*winsPerDay+start_value
                  index = self.model.index(0,i)
                  self.model.setData(index, todaysStarsTarget)
                  self.estStarsBox.clear()
                  if i == 6:
                        self.estStarsBox.setPlainText(str(end_value))
            self.tournamentTable.show()
            self.saveStarsTableToCSV(os.path.join('_internal','starstable.csv'))
class CrewTrainerDialogBox(QtWidgets.QDialog):
      def __init__(self):
            super().__init__()
            uic.loadUi(os.path.join('_internal','pss-ttt-crewtrainer.ui'), self)

            self.trainingList = [
                  ("ABL Green", "Steam Yoga", [0,0,4,1,0,0,0,0,0]),
                  ("ABL Blue", "Crew vs Wild", [1,1,8,1,0,0,0,0,0]),
                  ("ABL Gold", "Space Marine", [1,1,12,2,0,0,0,0,0]),
                  ("ABL Common", "Paracetamol", [1,0,2,1,0,1,0,0,0]),
                  ("ABL Elite", "Paracetamol Rapid", [2,0,4,1,0,2,0,0,0]),
                  ("ABL Unique", "Ibuprofen", [3,0,8,3,0,3,0,0,1]),
                  ("ABL Epic", "Ibuprofen Rapid", [5,0,16,3,0,5,0,0,2]),
                  ("ABL Hero", "Ginkgo", [12,0,24,4,0,12,0,0,3]),
                  ("ABL Special", "Brain Enhancer", [8,0,48,3,0,8,0,0,2]),
                  ("ABL Legendary", "Super Brain Enhancer", [3,0,96,1,0,3,0,0,1]),
                  ("ATK Green", "Kickbox", [0,4,0,1,0,0,0,0,0]),
                  ("ATK Blue", "BJJ", [1,8,1,1,0,0,0,0,0]),
                  ("ATK Gold", "Shaolin Tradition", [1,12,1,2,0,0,0,0,0]),
                  ("ATK Common", "Chicken Skewer", [0,2,1,1,0,0,0,0,1]),
                  ("ATK Elite", "Yakitori", [0,4,2,1,0,0,0,0,2]),
                  ("ATK Unique", "Double Yakitori", [1,8,3,3,0,0,0,0,3]),
                  ("ATK Epic", "Shish Kebabs", [2,16,5,3,0,0,0,0,5]),
                  ("ATK Hero", "Drumsticks", [3,24,12,4,0,0,0,0,12]),
                  ("ATK Special", "Steak", [2,48,8,3,0,0,0,0,8]),
                  ("ATK Legendary", "Roast Turkey", [1,96,3,1,0,0,0,0,3]),
                  ("ENG Green", "Study Expert Engineering Manual", [0,0,0,0,1,0,0,4,0]),
                  ("ENG Blue", "Engineering Summit", [0,0,0,0,1,1,1,8,0]),
                  ("ENG Gold", "Engineering PhD", [0,0,0,0,2,0,0,12,1]),
                  ("ENG Common", "Standard Engineering Tool Kit", [0,0,0,1,1,0,1,2,0]),
                  ("ENG Elite", "Obsolete Engineering Tool Kit", [0,0,0,1,2,0,2,4,0]),
                  ("ENG Unique", "Starter Engineering Tool Kit", [0,0,0,3,3,1,3,8,0]),
                  ("ENG Epic", "Advanced Engineering Tool Kit", [0,0,0,3,5,2,5,16,0]),
                  ("ENG Hero", "Rare Engineering Tool Kit", [0,0,0,4,12,3,12,24,0]),
                  ("ENG Special", "Prototype Engineering Tool Kit", [0,0,0,3,8,2,8,48,0]),
                  ("ENG Legendary", "Alien Engineering Tool Kit", [0,0,0,1,3,1,3,96,0]),
                  ("HP Green", "Bench Press", [4,0,0,1,0,0,0,0,0]),
                  ("HP Blue", "Muscle Beach", [8,1,1,1,0,0,0,0,0]),
                  ("HP Gold", "Olympic Weightlifting", [12,1,1,2,0,0,0,0,0]),
                  ("HP Common", "Small Protein Shake", [2,1,0,1,0,0,0,1,0]),
                  ("HP Elite", "Regular Protein Shake", [4,2,0,1,0,0,0,2,0]),
                  ("HP Unique", "Large Protein Shake", [8,3,1,3,0,0,0,3,0]),
                  ("HP Epic", "Super Protein Shake", [16,5,2,3,0,0,0,5,0]),
                  ("HP Hero", "HGH", [24,12,3,4,0,0,0,12,0]),
                  ("HP Special", "Enhanced HGH", [48,8,2,3,0,0,0,8,0]),
                  ("HP Legendary", "Prototype HGH", [96,3,1,1,0,0,0,3,0]),
                  ("PLT Green", "Read Expert Pilot Handbook", [0,0,0,0,0,4,0,0,1]),
                  ("PLT Blue", "Pilot Summit", [0,0,0,0,1,8,0,1,1]),
                  ("PLT Gold", "Pilot Expert", [0,0,0,0,1,12,0,0,2]),
                  ("PLT Common", "Street Map", [0,0,0,1,0,2,0,1,1]),
                  ("PLT Elite", "Travel Map", [0,0,0,1,0,4,0,2,2]),
                  ("PLT Unique", "World Map", [0,1,0,3,0,8,0,3,3]),
                  ("PLT Epic", "Global Map", [0,2,0,3,0,16,0,5,5]),
                  ("PLT Hero", "System Map", [0,3,0,4,0,24,0,12,12]),
                  ("PLT Special", "Star Map", [0,2,0,3,0,48,0,8,8]),
                  ("PLT Legendary", "Galactic Navigation Map", [0,1,0,1,0,96,0,3,3]),
                  ("RPR Common", "Repair Guide", [0,0,0,1,2,0,1,1,0]),
                  ("RPR Elite", "New Repair Guide", [0,0,0,1,4,0,2,2,0]),
                  ("RPR Unique", "Advanced Repair Guide", [0,0,1,3,8,0,3,3,0]),
                  ("RPR Epic", "Epic Repair Guide", [0,0,2,3,16,0,5,5,0]),
                  ("RPR Hero", "Lost Repair Guide", [0,0,3,4,24,0,12,12,0]),
                  ("RPR Special", "Special Repair Guide", [0,0,2,3,48,0,8,8,0]),
                  ("RPR Legendary", "Legendary Repair Guide", [0,0,1,1,96,0,3,3,0]),
                  ("SCI Green", "Big Book of Science", [0,0,0,0,0,0,4,1,0]),
                  ("SCI Blue", "Scientific Summit", [0,0,0,0,1,0,8,1,1]),
                  ("SCI Gold", "Science PhD", [0,0,0,0,0,0,12,2,1]),
                  ("SCI Common", "Drop of Brain Juice", [0,0,1,1,1,0,2,0,0]),
                  ("SCI Elite", "Sliver of Brain Juice", [0,0,2,1,2,0,4,0,0]),
                  ("SCI Unique", "Brew of Brain Juice", [0,0,3,3,3,0,8,0,1]),
                  ("SCI Epic", "Tincture of Brain Juice", [0,0,5,3,5,0,16,0,2]),
                  ("SCI Hero", "Solution of Brain Juice", [0,0,12,4,12,0,24,0,3]),
                  ("SCI Special", "Concentrate of Brain Juice", [0,0,8,3,8,0,48,0,2]),
                  ("SCI Legendary", "Elixir of Brain Juice", [0,0,3,1,3,0,96,0,1]),
                  ("STA Green", "Weighted Run", [0,0,0,5,0,0,0,0,0]),
                  ("STA Blue", "Hardcore Aerobics", [2,2,2,8,0,0,0,0,0]),
                  ("STA Gold", "Everest Climb", [3,3,3,16,0,0,0,0,0]),
                  ("STA Common", "Small Cola", [1,1,1,2,1,1,1,1,1]),
                  ("STA Elite", "Large Cola", [1,1,1,4,1,1,1,1,1]),
                  ("STA Unique", "Mountain Brew", [2,2,2,8,2,2,2,2,2]),
                  ("STA Epic", "Pink Cow", [3,3,3,16,3,3,3,3,3]),
                  ("STA Hero", "Large Pink Cow", [4,4,4,24,4,4,4,4,4]),
                  ("STA Special", "U", [3,3,3,48,3,3,3,3,3]),
                  ("STA Legendary", "Father", [2,2,2,96,2,2,2,2,2]),
                  ("WPN Green", "Read Expert Weapon Theory", [0,0,0,0,0,0,1,0,4]),
                  ("WPN Blue", "Weapons Summit", [0,0,0,0,1,1,1,0,8]),
                  ("WPN Gold", "Weapons PhD", [0,0,0,0,2,0,0,1,12]),
                  ("WPN Common", "Military Recruit Handbook", [0,1,0,1,0,1,0,0,2]),
                  ("WPN Elite", "Standard Combat Manual", [0,2,0,1,0,2,0,0,4]),
                  ("WPN Unique", "Galetrooper Training Manual", [0,3,0,3,0,3,1,0,8]),
                  ("WPN Epic", "Elite Combat Manual", [0,5,0,3,0,5,2,0,16]),
                  ("WPN Hero", "Veteran's Guidebook", [0,12,0,4,0,12,3,0,24]),
                  ("WPN Special", "How To Shoot Your Shot'", [0,8,0,3,0,8,2,0,48]),
                  ("WPN Legendary", "Sharpshooter's Cheatbook", [0,3,0,1,0,3,1,0,96]),
            ]
            self.trainingChart = [
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0]),
                  (" ", " ", [0,0,0,0,0,0,0,0,0])
            ]
            self.crewStats = [[0, 0] for _ in range(9)]
            self.trainingStats = [[0] for _ in range(9)]
            self.itemTrainingStats = [[0] for _ in range(9)]
            self.statsTable = self.findChild(QTableView, "crewStatTable")
            self.model = self.StatsTableModel(self.crewStats, self)
            self.chartTable = self.findChild(QTableView, "trainingChartTable")
            self.chartModel = self.TrainingChartTableModel(self.trainingChart, self.trainingStatBox)
            self.trainingStatBox = self.findChild(QtWidgets.QComboBox, "trainingStatBox")
            self.statsTable.setModel(self.model)
            self.statsTable.setColumnWidth(0,50)
            self.statsTable.setColumnWidth(1,90)
            self.statsTable.setColumnWidth(2,90)
            self.chartTable.setModel(self.chartModel)
            self.trainingPointsBox.currentIndexChanged.connect(self.onComboBoxValueChanged)
            self.trainingStatBox.currentIndexChanged.connect(self.onComboBoxValueChanged)
            self.trainingLevelBox.currentIndexChanged.connect(self.onComboBoxValueChanged)
            self.fatigueBox.currentIndexChanged.connect(self.onComboBoxValueChanged)

            self.presetsDialog = self.CrewTrainerPresetDialog(self.crewStats, parent=self)

            self.openPresetsButton.clicked.connect(self.open_presetWindow)
            
            for i in range(10):
                  if i == 0:
                        self.chartTable.setColumnWidth(i,125)
                  else:
                        self.chartTable.setColumnWidth(i,1)
            self.testPushButton.clicked.connect(self.wipeCrewStats)
            self.onComboBoxValueChanged()
      def open_presetWindow(self):
            self.presetsDialog.exec()
      def calculateTrainingChart(self):
            selected_item = self.trainingStatBox.currentText()
            selected_key = selected_item[:3]
            selected_data_list = [(name, data) for key, name, data in self.trainingList if key.startswith(selected_key)]
            self.model = self.TrainingChartTableModel(selected_data_list, self.trainingStatBox)
            self.chartTable.setModel(self.model)
      def updateFatigueMod(self):
            max_training_points = int(self.trainingPointsBox.currentText())
            total_tp = sum(int(self.crewStats[i][0]) for i in range (9))
            fatigue = self.fatigueBox.currentText()
            fatigue_m = {
                  '0': 1,
                  '1-50': 0.5,
                  '51-99': 0.3,
                  '100': 0
            }
            fatigue_m = fatigue_m.get(fatigue, 1)
            crew_stats_m = {}
            stat_names = [('hp', 1), ('atk', 1), ('abl', 1), ('sta', 1), ('rpr', 1), ('plt', 1), ('sci', 1), ('eng', 1), ('wpn', 1)]
            for i, (stat_name, _) in enumerate(stat_names):
                  crew_stat_value = self.crewStats[i][0]
                  crew_stats_m[stat_name] = float(fatigue_m) * (1 - (total_tp / max_training_points)) * (1 - (float(crew_stat_value) / max_training_points))
            for i, (stat_name, _) in enumerate(stat_names):
                  self.crewStats[i][1] = round(crew_stats_m[stat_name],3)
            self.statsTable.viewport().update()
            return
      def modifyTrainingMethods(self): #Added minimum values for final 3 cons of each type
            selected_training_stat = self.trainingStatBox.currentText()
            modified_data_list = []
            header_labels = self.getHeaderLabels(self.chartTable)
            default_limits = {}
            
            for item in self.trainingList:
                  key, name, data = item[0], item[1], item[2]
                  if key.startswith(selected_training_stat[:3]):
                        modified_data_list.append((name, key, data))

            self.resetTrainingData()
            if selected_training_stat == "RPR":
                  default_limits = {6: 4, 5: 2, 4: 1}
            else:
                  default_limits = {9: 4, 8: 2, 7: 1}

            num_rows = 7 if selected_training_stat == "RPR" else 10
            for i in range(num_rows):
                  for j in range(9):
                        modified_data_list[i][2][j] *= self.crewStats[j][1]
                        modified_data_list[i][2][j] = math.floor(modified_data_list[i][2][j])
                        if i in default_limits and modified_data_list[i][2][j] <= default_limits[i]:
                              if selected_training_stat == header_labels[j+1]:
                                    modified_data_list[i][2][j] = default_limits[i]
      def getHeaderLabels(self, tableView):
            model = tableView.model()
            if model is None:
                  print("No model set for the table view.")
                  return []
            header_labels = []
            for column in range(model.columnCount(0)):
                  header_label = model.headerData(column, Qt.Orientation.Horizontal)
                  header_labels.append(header_label)
            return header_labels
      def onComboBoxValueChanged(self):
            self.updateFatigueMod()
            self.calculateTrainingChart()
            self.modifyTrainingMethods()
      def resetTrainingData(self):
            self.trainingList = [
                  ("ABL Green", "Steam Yoga", [0,0,4,1,0,0,0,0,0]),
                  ("ABL Blue", "Crew vs Wild", [1,1,8,1,0,0,0,0,0]),
                  ("ABL Gold", "Space Marine", [1,1,12,2,0,0,0,0,0]),
                  ("ABL Common", "Paracetamol", [1,0,2,1,0,1,0,0,0]),
                  ("ABL Elite", "Paracetamol Rapid", [2,0,4,1,0,2,0,0,0]),
                  ("ABL Unique", "Ibuprofen", [3,0,8,3,0,3,0,0,1]),
                  ("ABL Epic", "Ibuprofen Rapid", [5,0,16,3,0,5,0,0,2]),
                  ("ABL Hero", "Ginkgo", [12,0,24,4,0,12,0,0,3]),
                  ("ABL Special", "Brain Enhancer", [8,0,48,3,0,8,0,0,2]),
                  ("ABL Legendary", "Super Brain Enhancer", [3,0,96,1,0,3,0,0,1]),
                  ("ATK Green", "Kickbox", [0,4,0,1,0,0,0,0,0]),
                  ("ATK Blue", "BJJ", [1,8,1,1,0,0,0,0,0]),
                  ("ATK Gold", "Shaolin Tradition", [1,12,1,2,0,0,0,0,0]),
                  ("ATK Common", "Chicken Skewer", [0,2,1,1,0,0,0,0,1]),
                  ("ATK Elite", "Yakitori", [0,4,2,1,0,0,0,0,2]),
                  ("ATK Unique", "Double Yakitori", [1,8,3,3,0,0,0,0,3]),
                  ("ATK Epic", "Shish Kebabs", [2,16,5,3,0,0,0,0,5]),
                  ("ATK Hero", "Drumsticks", [3,24,12,4,0,0,0,0,12]),
                  ("ATK Special", "Steak", [2,48,8,3,0,0,0,0,8]),
                  ("ATK Legendary", "Roast Turkey", [1,96,3,1,0,0,0,0,3]),
                  ("ENG Green", "Study Expert Engineering Manual", [0,0,0,0,1,0,0,4,0]),
                  ("ENG Blue", "Engineering Summit", [0,0,0,0,1,1,1,8,0]),
                  ("ENG Gold", "Engineering PhD", [0,0,0,0,2,0,0,12,1]),
                  ("ENG Common", "Standard Engineering Tool Kit", [0,0,0,1,1,0,1,2,0]),
                  ("ENG Elite", "Obsolete Engineering Tool Kit", [0,0,0,1,2,0,2,4,0]),
                  ("ENG Unique", "Starter Engineering Tool Kit", [0,0,0,3,3,1,3,8,0]),
                  ("ENG Epic", "Advanced Engineering Tool Kit", [0,0,0,3,5,2,5,16,0]),
                  ("ENG Hero", "Rare Engineering Tool Kit", [0,0,0,4,12,3,12,24,0]),
                  ("ENG Special", "Prototype Engineering Tool Kit", [0,0,0,3,8,2,8,48,0]),
                  ("ENG Legendary", "Alien Engineering Tool Kit", [0,0,0,1,3,1,3,96,0]),
                  ("HP Green", "Bench Press", [4,0,0,1,0,0,0,0,0]),
                  ("HP Blue", "Muscle Beach", [8,1,1,1,0,0,0,0,0]),
                  ("HP Gold", "Olympic Weightlifting", [12,1,1,2,0,0,0,0,0]),
                  ("HP Common", "Small Protein Shake", [2,1,0,1,0,0,0,1,0]),
                  ("HP Elite", "Regular Protein Shake", [4,2,0,1,0,0,0,2,0]),
                  ("HP Unique", "Large Protein Shake", [8,3,1,3,0,0,0,3,0]),
                  ("HP Epic", "Super Protein Shake", [16,5,2,3,0,0,0,5,0]),
                  ("HP Hero", "HGH", [24,12,3,4,0,0,0,12,0]),
                  ("HP Special", "Enhanced HGH", [48,8,2,3,0,0,0,8,0]),
                  ("HP Legendary", "Prototype HGH", [96,3,1,1,0,0,0,3,0]),
                  ("PLT Green", "Read Expert Pilot Handbook", [0,0,0,0,0,4,0,0,1]),
                  ("PLT Blue", "Pilot Summit", [0,0,0,0,1,8,0,1,1]),
                  ("PLT Gold", "Pilot Expert", [0,0,0,0,1,12,0,0,2]),
                  ("PLT Common", "Street Map", [0,0,0,1,0,2,0,1,1]),
                  ("PLT Elite", "Travel Map", [0,0,0,1,0,4,0,2,2]),
                  ("PLT Unique", "World Map", [0,1,0,3,0,8,0,3,3]),
                  ("PLT Epic", "Global Map", [0,2,0,3,0,16,0,5,5]),
                  ("PLT Hero", "System Map", [0,3,0,4,0,24,0,12,12]),
                  ("PLT Special", "Star Map", [0,2,0,3,0,48,0,8,8]),
                  ("PLT Legendary", "Galactic Navigation Map", [0,1,0,1,0,96,0,3,3]),
                  ("RPR Common", "Repair Guide", [0,0,0,1,2,0,1,1,0]),
                  ("RPR Elite", "New Repair Guide", [0,0,0,1,4,0,2,2,0]),
                  ("RPR Unique", "Advanced Repair Guide", [0,0,1,3,8,0,3,3,0]),
                  ("RPR Epic", "Epic Repair Guide", [0,0,2,3,16,0,5,5,0]),
                  ("RPR Hero", "Lost Repair Guide", [0,0,3,4,24,0,12,12,0]),
                  ("RPR Special", "Special Repair Guide", [0,0,2,3,48,0,8,8,0]),
                  ("RPR Legendary", "Legendary Repair Guide", [0,0,1,1,96,0,3,3,0]),
                  ("SCI Green", "Big Book of Science", [0,0,0,0,0,0,4,1,0]),
                  ("SCI Blue", "Scientific Summit", [0,0,0,0,1,0,8,1,1]),
                  ("SCI Gold", "Science PhD", [0,0,0,0,0,0,12,2,1]),
                  ("SCI Common", "Drop of Brain Juice", [0,0,1,1,1,0,2,0,0]),
                  ("SCI Elite", "Sliver of Brain Juice", [0,0,2,1,2,0,4,0,0]),
                  ("SCI Unique", "Brew of Brain Juice", [0,0,3,3,3,0,8,0,1]),
                  ("SCI Epic", "Tincture of Brain Juice", [0,0,5,3,5,0,16,0,2]),
                  ("SCI Hero", "Solution of Brain Juice", [0,0,12,4,12,0,24,0,3]),
                  ("SCI Special", "Concentrate of Brain Juice", [0,0,8,3,8,0,48,0,2]),
                  ("SCI Legendary", "Elixir of Brain Juice", [0,0,3,1,3,0,96,0,1]),
                  ("STA Green", "Weighted Run", [0,0,0,5,0,0,0,0,0]),
                  ("STA Blue", "Hardcore Aerobics", [2,2,2,8,0,0,0,0,0]),
                  ("STA Gold", "Everest Climb", [3,3,3,16,0,0,0,0,0]),
                  ("STA Common", "Small Cola", [1,1,1,2,1,1,1,1,1]),
                  ("STA Elite", "Large Cola", [1,1,1,4,1,1,1,1,1]),
                  ("STA Unique", "Mountain Brew", [2,2,2,8,2,2,2,2,2]),
                  ("STA Epic", "Pink Cow", [3,3,3,16,3,3,3,3,3]),
                  ("STA Hero", "Large Pink Cow", [4,4,4,24,4,4,4,4,4]),
                  ("STA Special", "U", [3,3,3,48,3,3,3,3,3]),
                  ("STA Legendary", "Father", [2,2,2,96,2,2,2,2,2]),
                  ("WPN Green", "Read Expert Weapon Theory", [0,0,0,0,0,0,1,0,4]),
                  ("WPN Blue", "Weapons Summit", [0,0,0,0,1,1,1,0,8]),
                  ("WPN Gold", "Weapons PhD", [0,0,0,0,2,0,0,1,12]),
                  ("WPN Common", "Military Recruit Handbook", [0,1,0,1,0,1,0,0,2]),
                  ("WPN Elite", "Standard Combat Manual", [0,2,0,1,0,2,0,0,4]),
                  ("WPN Unique", "Galetrooper Training Manual", [0,3,0,3,0,3,1,0,8]),
                  ("WPN Epic", "Elite Combat Manual", [0,5,0,3,0,5,2,0,16]),
                  ("WPN Hero", "Veteran's Guidebook", [0,12,0,4,0,12,3,0,24]),
                  ("WPN Special", "How To Shoot Your Shot'", [0,8,0,3,0,8,2,0,48]),
                  ("WPN Legendary", "Sharpshooter's Cheatbook", [0,3,0,1,0,3,1,0,96]),
            ]
      def wipeCrewStats(self):
            for i in range(9):
                  self.crewStats[i][0] = 0
            self.statsTable.viewport().update()
            self.onComboBoxValueChanged()
      class TrainingStatTableModel(QAbstractTableModel):
            def __init__(self, values, parent=None):
                  super().__init__()
                  self.values = values
                  self.verticalHeaders = ['HP','ATK','ABL','STA','RPR','PLT','SCI','ENG','WPN']
                  self.horizontalHeaders = ['']
            def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
                  if role == Qt.ItemDataRole.DisplayRole:
                        if orientation == Qt.Orientation.Horizontal and section < len(self.horizontalHeaders):
                              return self.horizontalHeaders[section]
                        elif role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Vertical:
                              return self.verticalHeaders[section]
                  return None
            def rowCount(self, parent):
                  return len(self.values)
            def columnCount(self, parent):
                  return 1  # Assuming you have only one column
            def data(self, index, role):
                  if role == Qt.ItemDataRole.DisplayRole:
                        row = index.row()
                        if 0 <= row < len(self.values):
                              return str(self.values[row])
                  return None
            def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
                  if index.isValid() and role == Qt.ItemDataRole.EditRole:
                        if isinstance(value, int):
                              self.values[index.row()] = value
                              self.dataChanged.emit(index, index)
                              self.parent.onComboBoxValueChanged()
                              return True
                        else:
                              return False
                  return False
      class TrainingChartTableModel(QAbstractTableModel):
            def __init__(self, data, trainingStatBox, parent=None):
                  super().__init__(parent)
                  self._data = data
                  self.trainingStatBox = trainingStatBox
                  self.horizontalHeaders = ['Training', 'HP','ATK','ABL','STA','RPR','PLT','SCI','ENG','WPN']
                  self.verticalHeaders = [''] * len(data)
            def rowCount(self, index):
                  return len(self._data)
            def columnCount(self, index):
                  return len(self.horizontalHeaders)
            def data(self, index, role=Qt.ItemDataRole.DisplayRole):
                  if role == Qt.ItemDataRole.DisplayRole:
                        row = index.row()
                        col = index.column()
                        if col == 0:
                              return self._data[row][0]  # First column data
                        elif 0 < col <= len(self.horizontalHeaders) - 1:
                              item_data = self._data[row][1]
                              if isinstance(item_data, str):
                                    return ""  # Return an empty string for non-existent data
                              else:
                                    return item_data[col - 1]  # Data values (adjusting for the first column)
                  elif role == Qt.ItemDataRole.BackgroundRole:
                        value = self.data(index, Qt.ItemDataRole.DisplayRole)
                        col = index.column()
                        if value == 0:
                              return QColor(198,239,206)
                        elif self.trainingStatBox.currentText() == self.horizontalHeaders[col]:
                              return QColor(189,215,238)
                        else:
                              return QColor(255,255,255)
                  return None
            def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
                  if role == Qt.ItemDataRole.EditRole and index.isValid():
                        row = index.row()
                        col = index.column()
                        self._data[row][2][col - 1] = value
                        self.dataChanged.emit(index, index)
                        return True
                  return False
            def updateData(self, data):
                  self.beginResetModel()
                  self._data = data
                  self.endResetModel()
            def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
                  if role == Qt.ItemDataRole.DisplayRole:
                        if orientation == Qt.Orientation.Horizontal:
                              if section == 0:
                                    return "Name"
                              else:
                                    return f"{self.horizontalHeaders[section]}"
                        else:
                              return str(section + 1)
                  return None
            def getCellValue(self, row, column):
                  return self._data[row][column]
            def flags(self, index):
                  return Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
      class StatsTableModel(QAbstractTableModel):
            def __init__(self, data, parent=None):
                  super().__init__()
                  self._data = data
                  self.parent = parent
                  self.verticalHeaders = ['HP','ATK','ABL','STA','RPR','PLT','SCI','ENG','WPN']
                  self.horizontalHeaders = ['Current', 'Multiplier', '% Required']
            def rowCount(self, index):
                  return len(self._data)
            def columnCount(self, index):
                  return len(self._data[0])
            def data(self, index, role=Qt.ItemDataRole.DisplayRole):
                  if index.isValid() and role == Qt.ItemDataRole.DisplayRole:
                        return str(self._data[index.row()][index.column()])
            def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
                  if index.isValid() and role == Qt.ItemDataRole.EditRole:
                        try:
                              int_value = int(value)
                        except ValueError:
                              int_value = 0
                        self._data[index.row()][index.column()] = int_value
                        self.dataChanged.emit(index, index)
                        self.parent.onComboBoxValueChanged()
                        return True
                  return False
            def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
                  if role == Qt.ItemDataRole.DisplayRole:
                        if orientation == Qt.Orientation.Horizontal and section < len(self.horizontalHeaders):
                              return self.horizontalHeaders[section]
                        elif role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Vertical:
                              return self.verticalHeaders[section]
                  return None
            def getCellValue(self, row, column):
                  return self._data[row][column]
            def flags(self, index):
                  if index.column() >= 1:
                        return Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
                  else:
                        return Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable
      class CrewTrainerPresetDialog(QtWidgets.QDialog):
            def __init__(self, crewStats, parent=None):
                  super().__init__(parent)
                  uic.loadUi(os.path.join('_internal','pss-ttt-ctp.ui'), self)
                  self.crewStats = crewStats
                  self.savePresetButton.clicked.connect(self.savePresetData)
                  self.loadPresetButton.clicked.connect(self.loadPresetData)
                  self.deletePresetButton.clicked.connect(self.deleteSelectedPreset)

                  self.model = QStandardItemModel()
                  self.presetTableView.setModel(self.model)
                  self.model.setHorizontalHeaderLabels(['Crew Name', 'HP','ATK','ABL','STA','RPR','PLT','SCI','ENG','WPN'])
                  for i in range(self.model.columnCount()):
                        if i==0:
                              self.presetTableView.setColumnWidth(i,100)
                        else:
                              self.presetTableView.setColumnWidth(i,20)
                  self.loadPresetCSV()
            def loadPresetData(self):
                  selected_row = self.presetTableView.currentIndex().row()
                  if selected_row != -1:
                        for i in range(1, self.model.columnCount()):
                              data = self.model.index(selected_row, i).data()
                              self.crewStats[i - 1][0] = data
                              self.crewStats[i - 1][1] = 1.0
                  self.accept()
                  self.parent().updateFatigueMod()
            def savePresetData(self):
                  first_column_data = [row[0] for row in self.crewStats]

                  # Getting the name from self.presetCrewNameBox
                  preset_name = self.presetCrewNameBox.toPlainText()

                  # Appending data to presetTableView
                  row_count = self.model.rowCount()
                  self.model.insertRow(row_count)
                  self.model.setData(self.model.index(row_count, 0), preset_name)
                  for i, data in enumerate(first_column_data):
                        self.model.setData(self.model.index(row_count, i + 1), data)
                  for i in range(self.model.columnCount()):
                        if i==0:
                              self.presetTableView.setColumnWidth(i,100)
                        else:
                              self.presetTableView.setColumnWidth(i,20)
                  self.savePresetCSV()
            def savePresetCSV(self):
                  with open(os.path.join('_internal','presets.csv'), 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        for i in range(self.model.rowCount()):
                              row_data = []
                              for j in range(self.model.columnCount()):
                                    data = self.model.index(i, j).data()
                                    row_data.append(data)
                              writer.writerow(row_data)
            def loadPresetCSV(self):
                  filename = os.path.join('_internal','presets.csv')
                  if os.path.exists(filename):
                        with open(filename, 'r', newline='', encoding='utf-8') as csvfile:
                              reader = csv.reader(csvfile)
                              self.model.removeRows(0, self.model.rowCount())
                              for row in reader:
                                    row_position = self.model.rowCount()
                                    self.model.insertRow(row_position)
                                    for column, value in enumerate(row):
                                          item = QtGui.QStandardItem(value)
                                          self.model.setItem(row_position, column, item)
                  else:
                        self.savePresetCSV()
            def deleteSelectedPreset(self):
                  selected_index = self.presetTableView.selectionModel().selectedRows()
                  for index in sorted(selected_index, reverse=True):
                        self.model.removeRow(index.row())
                  self.savePresetCSV()
class StarTargetTrackDialogBox(QtWidgets.QDialog):
      copyStarsTargetTrackClicked = QtCore.pyqtSignal(str, bool)
      currentStarsTarget = ' '
      def __init__(self, parent=None):
            super().__init__(parent)
            uic.loadUi(os.path.join('_internal','pss-ttt-stt.ui'), self)

            self.table_widgets = {
            "dayFour": {"button_add": self.dayFourAdd, "button_remove": self.dayFourRemove, "model": QtGui.QStandardItemModel()},
            "dayFive": {"button_add": self.dayFiveAdd, "button_remove": self.dayFiveRemove, "model": QtGui.QStandardItemModel()},
            "daySix": {"button_add": self.daySixAdd, "button_remove": self.daySixRemove, "model": QtGui.QStandardItemModel()},
            "daySeven": {"button_add": self.daySevenAdd, "button_remove": self.daySevenRemove, "model": QtGui.QStandardItemModel()}
            }

            for day, widgets in self.table_widgets.items():
                  table = getattr(self, f"{day}Table")
                  widgets["model"].setHorizontalHeaderLabels(['Player Name', 'Fleet Name', 'Stars', 'Max Trophy'])
                  table.setModel(widgets["model"])
                  widgets["button_add"].clicked.connect(lambda _, day=day: self.addToTargetTable(day))
                  widgets["button_remove"].clicked.connect(lambda _, day=day: self.removeTargetTable(day))
            self.loadStarsCSV()
            self.trackerResetButton.clicked.connect(self.resetStarsTargetList)
            for day in ["dayFour", "dayFive", "daySix", "daySeven"]:
                        getattr(self, f"{day}Copy").clicked.connect(lambda _, day=day: self.copyDayToSearchBox(day))
      def copyDayToSearchBox(self, day):
            table = getattr(self, f"{day}Table")
            selected_indexes = table.selectionModel().selectedRows()
            if not selected_indexes:
                  throwErrorMessage("No row selected", "Please select a row")
                  return

            selected_row = selected_indexes[0].row()            
            player_name = table.model().index(selected_row, 0).data()
            parent_widget = self.parent()
            if parent_widget:
                  parent_widget.playerNameSearchBox.setPlainText(player_name)
                  parent_widget.searchPlayer()
                  self.close()
            else:
                  throwErrorMessage("Error finding Main Window", "Close and reopen the application fully")
      def keyPressEvent(self, event):
            if event.key() == QtCore.Qt.Key.Key_Escape:
                  self.close()
            else:
                  super().keyPressEvent(event)
      def closeEvent(self, event):
            self.saveStarsCSV()
            super().closeEvent(event)
      def addToTargetTable(self, day):
            player_name = self.playerNameSTTBox.toPlainText()
            fleet_name = self.fleetNameSTTBox.toPlainText()
            last_stars = self.lastStarsSTTBox.toPlainText()
            max_trophies = self.maxTrophiesSTTBox.toPlainText()

            target_table = getattr(self, f"{day}Table")
            target_model = target_table.model()

            if target_model is None:
                  throwErrorMessage("Error", traceback.extract_stack()[-2].lineno)
            new_row = [QtGui.QStandardItem(player_name), QtGui.QStandardItem(fleet_name), QtGui.QStandardItem(last_stars), QtGui.QStandardItem(max_trophies)]

            target_model = target_table.model()
            target_table.setColumnWidth(1,100)
            target_table.setColumnWidth(2,50)
            target_table.setColumnWidth(3,75)
            if target_model is not None:
                  target_model.appendRow(new_row)
            else:
                  print("No model set for the target table")
            self.saveStarsCSV()
      def populateSTT(self, player_name):
            self.playerNameSTTBox.setPlainText(player_name)
            if not QSqlDatabase.database("targetdb").isOpen():
                  if not QSqlDatabase.database("targetdb").open():
                        throwErrorMessage("Database Connection Failure", "Targets DB did not open properly")
                        return
            query = QSqlQuery(QSqlDatabase.database("targetdb"))
            query.prepare("SELECT playername, fleetname, laststars, maxtrophies FROM players WHERE playername = :player_name")
            query.bindValue(":player_name", player_name)
            if not query.exec():
                  throwErrorMessage("Database Query Error", query.lastError().text())
                  return
    
            if query.next():
                  fleet_name, last_stars, max_trophies = query.value(1), query.value(2), query.value(3)
                  self.fleetNameSTTBox.setPlainText(fleet_name)
                  self.lastStarsSTTBox.setPlainText(str(last_stars))
                  self.maxTrophiesSTTBox.setPlainText(str(max_trophies))
            else:
                  self.fleetNameSTTBox.setPlainText(" ")
                  self.lastStarsSTTBox.setPlainText(" ")
                  self.maxTrophiesSTTBox.setPlainText(" ")
                  self.playerNameSTTBox.setPlainText(" ")
      def removeTargetTable(self, day):
            target_table = getattr(self, f"{day}Table")
            selected_indexes = target_table.selectionModel().selectedRows()
            for index in selected_indexes:
                  target_table.model().removeRow(index.row())
            self.saveStarsCSV()
      def resetStarsTargetList(self):
            confirmation_dialog = self.StarsTargetTrackConfirmationDialog(self)
            confirmation_dialog.exec()
      def saveStarsCSV(self):
            with open(os.path.join('_internal','starstrack.csv'), 'w', newline='', encoding='utf-8') as csvfile:
                  writer = csv.writer(csvfile)
                  # Write data for each table
                  for day, widgets in self.table_widgets.items():
                        model = widgets["model"]
                        for row in range(model.rowCount()):
                              data = [model.index(row, col).data() for col in range(model.columnCount())]
                              writer.writerow([day] + data)
      def loadStarsCSV(self):
            try:
                  with open(os.path.join('_internal','starstrack.csv'), 'r', newline='', encoding='utf-8') as csvfile:
                        reader = csv.reader(csvfile)
                        # Read data for each table
                        for row in reader:
                              day = row[0]
                              data = row[1:]
                              widgets = self.table_widgets[day]
                              model = widgets["model"]
                              model.appendRow([QtGui.QStandardItem(item) for item in data])
                              table = getattr(self, f"{day}Table")
                              table.setColumnWidth(0, 90)
                              table.setColumnWidth(1, 100)
                              table.setColumnWidth(2, 25)
                              table.setColumnWidth(3, 75)
            except FileNotFoundError:
                  print("CSV file not found, no data loaded.")
      class StarsTargetTrackConfirmationDialog(QtWidgets.QDialog):
            def __init__(self, parent=None):
                  super().__init__(parent)
                  uic.loadUi(os.path.join('_internal','pss-ttt-stt-confirm.ui'), self)

                  self.cancelButton.clicked.connect(self.reject)
                  self.confirmButton.clicked.connect(self.accept)
            def accept(self):
                  super().accept()

                  if self.parent():
                        self.parent().dayFourTable.model().clear()
                        self.parent().dayFiveTable.model().clear()
                        self.parent().daySixTable.model().clear()
                        self.parent().daySevenTable.model().clear()
                  else:
                        print("Parent not found")
            def reject(self):
                  super().reject()
      ## Startup and program operation
if __name__ == "__main__":
      app = QtWidgets.QApplication(sys.argv)
      if create_connection():
            create_table()
      else:
            sys.exit(1)
      window = MainWindow()
      app.exec()


''' DEFUNCT - but saving?
class PlayerDialogBox(QtWidgets.QDialog):
      copyPlayerSearchClicked = QtCore.pyqtSignal(str, bool)
      def __init__(self):
            super().__init__()
            uic.loadUi(os.path.join('_internal', 'pss-ttt-psb.ui'), self)
            self.playerSearchClose.clicked.connect(self.accept)
            self.copyPlayerSearch.clicked.connect(self.onCopyPlayerSearchClicked)
            self.playerFilterBox.textChanged.connect(self.filterPlayerList)
            self.playerCharLimiter.valueChanged.connect(self.filterPlayerListLength)
            self.playerNameCharLimit.stateChanged.connect(self.filterPlayerListLength)
      def populatePlayerList(self):
            self.playerNameList.clear()
            if not QSqlDatabase.database("targetdb").isOpen():
                  if not QSqlDatabase.database("targetdb").open():
                        throwErrorMessage("Database Connection Failure", "Targets DB did not open properly")
                        return
            
            query = QSqlQuery(QSqlDatabase.database("targetdb"))
            query.prepare("SELECT playername FROM players")
            if not query.exec():
                  throwErrorMessage("Query Execution Failre", query.lastError().text())
                  return
            while query.next():
                  player_name = query.value(0)
                  item = QListWidgetItem(player_name)
                  self.playerNameList.addItem(item)
      def onCopyPlayerSearchClicked(self):
            selected_item = self.playerNameList.currentItem()
            if selected_item:
                  selected_text = selected_item.text()
                  self.copyPlayerSearchClicked.emit(selected_text, True)
      def filterPlayerList(self):
            filter_text = self.playerFilterBox.toPlainText().lower()
            for i in range(self.playerNameList.count()):
                  item = self.playerNameList.item(i)
                  item_text = item.text().lower()
                  if filter_text in item_text:
                        item.setHidden(False)
                  else:
                        item.setHidden(True)
      def filterPlayerListLength(self):
            if self.playerNameCharLimit.isChecked():
                  length_threshhold = self.playerCharLimiter.value()
                  for i in range(self.playerNameList.count()):
                        item = self.playerNameList.item(i)
                        item_text = item.text()
                        if len(item_text) == length_threshhold:
                              item.setHidden(False)
                        else:
                              item.setHidden(True)
class FleetDialogBox(QtWidgets.QDialog):
      copyFleetSearchClicked = QtCore.pyqtSignal(str, bool)

      def __init__(self):
            super().__init__()
            uic.loadUi(os.path.join('_internal','pss-ttt-fleetbrowser.ui'), self)
            self.fleetSearchClose.clicked.connect(self.accept)
            self.copyFleetSearch.clicked.connect(self.onCopyFleetSearchClicked)
            self.fleetFilterBox.textChanged.connect(self.filterFleetList)
            self.charLimiter.valueChanged.connect(self.filterFleetListLength)
            self.charLimiterCheckBox.stateChanged.connect(self.filterFleetListLength)
      def populateFleetList(self, fleet_name):
            self.fleetNameList.clear()
            if not QSqlDatabase.database("targetdb").isOpen():
                  if not QSqlDatabase.database("targetdb").open():
                        throwErrorMessage("Database Connection Failure", "Targets DB did not open properly")
                        return

            query = QSqlQuery(QSqlDatabase.database("targetdb"))
            query.prepare("SELECT playername FROM players WHERE fleetname = :fleet_name")
            query.bindValue(":fleet_name", fleet_name)
            if not query.exec():
                  throwErrorMessage("Query Execution Failure", query.lastError().text())
                  return
            while query.next():
                  player_name = query.value(0)
                  item = QListWidgetItem(player_name)
                  self.fleetNameList.addItem(item)
      def onCopyFleetSearchClicked(self):
            selected_item = self.fleetNameList.currentItem()
            if selected_item:
                  selected_text = selected_item.text()
                  self.copyFleetSearchClicked.emit(selected_text, True)
      def filterFleetList(self):
            filter_text = self.fleetFilterBox.toPlainText().lower()
            for i in range(self.fleetNameList.count()):
                  item = self.fleetNameList.item(i)
                  item_text = item.text().lower()
                  if filter_text in item_text:
                        item.setHidden(False)
                  else:
                        item.setHidden(True)
      def filterFleetListLength(self):
            if self.charLimiterCheckBox.isChecked():
                  length_threshold = self.charLimiter.value()
                  for i in range(self.fleetNameList.count()):
                        item = self.fleetNameList.item(i)
                        item_text = item.text()
                        if len(item_text) == length_threshold:
                              item.setHidden(False)
                        else:
                              item.setHidden(True)
class FleetNameDialogBox(QtWidgets.QDialog):
      copyFleetNameSearchClicked = QtCore.pyqtSignal(str, bool)
      def __init__(self):
            super().__init__()
            uic.loadUi(os.path.join('_internal', 'pss-ttt-fnb.ui'), self)
            self.fleetNameSearchClose.clicked.connect(self.accept)
            self.copyFleetNameSearch.clicked.connect(self.onCopyFleetNameSearchClicked)
            self.fleetNameFilterBox.textChanged.connect(self.filterFleetNameList)
            self.fleetNameCharLimiter.valueChanged.connect(self.filterFleetNameListLength)
            self.fleetNameCharLimitCheckBox.stateChanged.connect(self.filterFleetNameListLength)
      def buildFleetList(self):
            self.fleetNameList.clear()
            if not QSqlDatabase.database("targetdb").isOpen():
                  if not QSqlDatabase.database("targetdb").open():
                        throwErrorMessage("Database Connection Failure", "Targets DB did not open properly")
                        return
            
            query = QSqlQuery(QSqlDatabase.database("targetdb"))
            query.prepare("SELECT fleetname FROM players")
            if not query.exec():
                  throwErrorMessage("Query Execution Failure", query.lastError().text())
                  return
            fleet_names = set()
            while query.next():
                  fleet_name = query.value(0)
                  if fleet_name not in fleet_names:
                        fleet_names.add(fleet_name)
                        item = QListWidgetItem(fleet_name)
                        self.fleetNameList.addItem(item)
      def onCopyFleetNameSearchClicked(self):
            selected_item = self.fleetNameList.currentItem()
            if selected_item:
                  selected_text = selected_item.text()
                  self.copyFleetNameSearchClicked.emit(selected_text, True)
      def filterFleetNameList(self):
            filter_text = self.fleetNameFilterBox.toPlainText().lower()
            for i in range(self.fleetNameList.count()):
                  item = self.fleetNameList.item(i)
                  item_text = item.text().lower()
                  if filter_text in item_text:
                        item.setHidden(False)
                  else:
                        item.setHidden(True)
      def filterFleetNameListLength(self):
            if self.fleetNameCharLimitCheckBox.isChecked():
                  length_threshhold = self.fleetNameCharLimiter.value()
                  for i in range(self.fleetNameList.count()):
                        item = self.fleetNameList.item(i)
                        item_text = item.text()
                        if len(item_text) == length_threshhold:
                              item.setHidden(False)
                        else:
                              item.setHidden(True)
'''